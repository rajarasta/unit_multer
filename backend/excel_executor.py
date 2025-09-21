"""
FastAPI Excel Executor - Agentic workflow backend for safe spreadsheet operations
"""

import asyncio
import json
import logging
import os
import tempfile
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
import openpyxl
from openpyxl.styles import Font, Fill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, HTTPException, UploadFile, File, Header, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
import uvicorn
from starlette.responses import Response

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global session store
sessions: Dict[str, Dict] = {}
SESSION_TTL = timedelta(hours=2)  # 2 hour session timeout

# Safe upload directory
UPLOAD_DIR = Path("uploads/excel")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# FastAPI app
app = FastAPI(title="Excel AI Agent API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# === Data Models ===

class CellStyle(BaseModel):
    backgroundColor: Optional[str] = None
    textColor: Optional[str] = None
    bold: Optional[bool] = None
    italic: Optional[bool] = None
    fontSize: Optional[int] = None
    horizontalAlign: Optional[str] = None  # left, center, right
    verticalAlign: Optional[str] = None    # top, center, bottom

class BaseAction(BaseModel):
    sheet: str
    target: str  # A1 notation

class UpdateCellAction(BaseAction):
    type: str = "updateCell"
    value: Union[str, int, float]

class FormatCellAction(BaseAction):
    type: str = "formatCell"
    style: CellStyle

class SetRowHeightAction(BaseAction):
    type: str = "setRowHeight"
    height: float

class SetColumnWidthAction(BaseAction):
    type: str = "setColumnWidth"
    width: float

class InsertRowAction(BaseAction):
    type: str = "insertRow"
    index: int
    count: int = 1

class DeleteRowAction(BaseAction):
    type: str = "deleteRow"
    start: int
    end: int

class InsertColumnAction(BaseAction):
    type: str = "insertColumn"
    index: int
    count: int = 1

class DeleteColumnAction(BaseAction):
    type: str = "deleteColumn"
    start: int
    end: int

class MergeCellsAction(BaseAction):
    type: str = "mergeCells"
    range: str  # A1:B2

class UnmergeCellsAction(BaseAction):
    type: str = "unmergeCells"
    range: str  # A1:B2

# Union of all action types
Action = Union[
    UpdateCellAction, FormatCellAction, SetRowHeightAction, SetColumnWidthAction,
    InsertRowAction, DeleteRowAction, InsertColumnAction, DeleteColumnAction,
    MergeCellsAction, UnmergeCellsAction
]

class ExecuteRequest(BaseModel):
    actions: List[Action]
    dryRun: bool = False
    transactionId: Optional[str] = None

class DiffItem(BaseModel):
    sheet: str
    row: int
    col: int
    before: Dict[str, Any]
    after: Dict[str, Any]

class ExecutionResult(BaseModel):
    applied: List[str]
    errors: List[str]
    diff: List[DiffItem]
    stats: Dict[str, int]
    traceId: str

class SheetInfo(BaseModel):
    name: str
    rows: int
    cols: int
    cellCount: int

class WorkbookInfo(BaseModel):
    name: str
    sheets: List[SheetInfo]
    totalCells: int

# === SSE Event Models ===

class SSEEvent(BaseModel):
    type: str  # status, finding, result, ask, error
    message: str
    data: Optional[Dict] = None
    timestamp: float = Field(default_factory=time.time)

# === Utility Functions ===

def create_session() -> str:
    """Create a new session with unique ID."""
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        "id": session_id,
        "created": datetime.now(),
        "workbook": None,
        "workbook_path": None,
        "events": [],
        "last_activity": datetime.now()
    }
    return session_id

def get_session(session_id: str) -> Dict:
    """Get session by ID, raise 404 if not found or expired."""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]

    # Check if session expired
    if datetime.now() - session["last_activity"] > SESSION_TTL:
        cleanup_session(session_id)
        raise HTTPException(status_code=404, detail="Session expired")

    # Update last activity
    session["last_activity"] = datetime.now()
    return session

def cleanup_session(session_id: str):
    """Clean up session and associated files."""
    if session_id in sessions:
        session = sessions[session_id]
        if session.get("workbook_path") and os.path.exists(session["workbook_path"]):
            os.remove(session["workbook_path"])
        del sessions[session_id]

def add_event(session_id: str, event: SSEEvent):
    """Add event to session for SSE streaming."""
    if session_id in sessions:
        sessions[session_id]["events"].append(event)

def style_to_openpyxl(style: CellStyle):
    """Convert CellStyle to openpyxl style objects."""
    font_kwargs = {}
    if style.bold is not None:
        font_kwargs["bold"] = style.bold
    if style.italic is not None:
        font_kwargs["italic"] = style.italic
    if style.fontSize is not None:
        font_kwargs["size"] = style.fontSize
    if style.textColor:
        font_kwargs["color"] = style.textColor.replace("#", "")

    fill = None
    if style.backgroundColor:
        fill = PatternFill(start_color=style.backgroundColor.replace("#", ""),
                          end_color=style.backgroundColor.replace("#", ""),
                          fill_type="solid")

    alignment = None
    if style.horizontalAlign or style.verticalAlign:
        alignment = Alignment(
            horizontal=style.horizontalAlign,
            vertical=style.verticalAlign
        )

    return {
        "font": Font(**font_kwargs) if font_kwargs else None,
        "fill": fill,
        "alignment": alignment
    }

def parse_a1_notation(a1: str) -> tuple:
    """Parse A1 notation to (row, col) - 1-indexed."""
    from openpyxl.utils import coordinate_from_string
    col, row = coordinate_from_string(a1)
    return row, openpyxl.utils.column_index_from_string(col)

# === API Endpoints ===

@app.post("/api/excel/session")
async def create_excel_session():
    """Create a new Excel session."""
    session_id = create_session()
    logger.info(f"Created Excel session: {session_id}")
    return {"sessionId": session_id}

@app.post("/api/excel/open")
async def open_excel_file(
    file: UploadFile = File(...),
    x_session_id: str = Header(..., alias="X-Session-Id")
):
    """Upload and open an Excel file in session."""
    session = get_session(x_session_id)

    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file format")

    try:
        # Save uploaded file
        file_path = UPLOAD_DIR / f"{x_session_id}_{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Load workbook
        workbook = openpyxl.load_workbook(file_path)
        session["workbook"] = workbook
        session["workbook_path"] = str(file_path)

        # Add event
        add_event(x_session_id, SSEEvent(
            type="status",
            message=f"Excel file '{file.filename}' loaded successfully"
        ))

        logger.info(f"Loaded Excel file: {file.filename} for session {x_session_id}")

        return {
            "workbook": {
                "name": file.filename,
                "sheets": len(workbook.sheetnames)
            }
        }

    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load file: {str(e)}")

@app.get("/api/excel/sheets")
async def get_sheets(sessionId: str):
    """Get list of sheets in workbook."""
    session = get_session(sessionId)
    workbook = session.get("workbook")

    if not workbook:
        raise HTTPException(status_code=400, detail="No workbook loaded")

    sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheets.append(SheetInfo(
            name=sheet_name,
            rows=sheet.max_row,
            cols=sheet.max_column,
            cellCount=sheet.max_row * sheet.max_column
        ))

    return {"sheets": sheets}

@app.get("/api/excel/range")
async def get_range(sessionId: str, sheet: str, range: str):
    """Get cell values and styles for a range."""
    session = get_session(sessionId)
    workbook = session.get("workbook")

    if not workbook:
        raise HTTPException(status_code=400, detail="No workbook loaded")

    if sheet not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet}' not found")

    ws = workbook[sheet]

    # Parse range (simple A1:B2 format)
    cells = []
    for row in ws[range]:
        row_data = []
        for cell in row:
            row_data.append({
                "value": cell.value,
                "coordinate": cell.coordinate
            })
        cells.append(row_data)

    return {"range": range, "cells": cells}

@app.post("/api/excel/ops")
async def execute_operations(
    request: ExecuteRequest,
    x_session_id: str = Header(..., alias="X-Session-Id"),
    background_tasks: BackgroundTasks = None
):
    """Execute spreadsheet operations."""
    session = get_session(x_session_id)
    workbook = session.get("workbook")

    if not workbook:
        raise HTTPException(status_code=400, detail="No workbook loaded")

    trace_id = request.transactionId or str(uuid.uuid4())

    # Add initial event
    add_event(x_session_id, SSEEvent(
        type="status",
        message=f"Starting execution of {len(request.actions)} operations..."
    ))

    applied = []
    errors = []
    diff = []

    try:
        for i, action in enumerate(request.actions):
            try:
                # Add progress event
                add_event(x_session_id, SSEEvent(
                    type="finding",
                    message=f"Executing {action.type} on {action.sheet}:{action.target}"
                ))

                # Get worksheet
                if action.sheet not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{action.sheet}' not found")

                ws = workbook[action.sheet]

                # Execute action based on type
                if action.type == "updateCell":
                    row, col = parse_a1_notation(action.target)
                    cell = ws.cell(row=row, column=col)
                    old_value = cell.value

                    if not request.dryRun:
                        cell.value = action.value

                    diff.append(DiffItem(
                        sheet=action.sheet,
                        row=row,
                        col=col,
                        before={"value": old_value},
                        after={"value": action.value}
                    ))

                elif action.type == "formatCell":
                    row, col = parse_a1_notation(action.target)
                    cell = ws.cell(row=row, column=col)

                    if not request.dryRun:
                        styles = style_to_openpyxl(action.style)
                        if styles["font"]:
                            cell.font = styles["font"]
                        if styles["fill"]:
                            cell.fill = styles["fill"]
                        if styles["alignment"]:
                            cell.alignment = styles["alignment"]

                    diff.append(DiffItem(
                        sheet=action.sheet,
                        row=row,
                        col=col,
                        before={"style": "original"},
                        after={"style": action.style.dict()}
                    ))

                elif action.type == "insertRow":
                    if not request.dryRun:
                        ws.insert_rows(action.index, action.count)

                    diff.append(DiffItem(
                        sheet=action.sheet,
                        row=action.index,
                        col=0,
                        before={"rows": ws.max_row - action.count},
                        after={"rows": ws.max_row}
                    ))

                elif action.type == "setRowHeight":
                    row, _ = parse_a1_notation(action.target)
                    if not request.dryRun:
                        ws.row_dimensions[row].height = action.height

                # Add more action types as needed...

                applied.append(f"{action.type} on {action.sheet}:{action.target}")

            except Exception as e:
                error_msg = f"Failed {action.type}: {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg)

        # Success event
        add_event(x_session_id, SSEEvent(
            type="result",
            message=f"Execution complete: {len(applied)} operations applied, {len(errors)} errors",
            data={"diff": diff}
        ))

        # Save workbook if not dry run
        if not request.dryRun and session.get("workbook_path"):
            workbook.save(session["workbook_path"])

        return ExecutionResult(
            applied=applied,
            errors=errors,
            diff=diff,
            stats={
                "totalActions": len(request.actions),
                "appliedActions": len(applied),
                "errorCount": len(errors),
                "cellsModified": len(diff)
            },
            traceId=trace_id
        )

    except Exception as e:
        error_msg = f"Execution failed: {str(e)}"
        add_event(x_session_id, SSEEvent(
            type="error",
            message=error_msg
        ))
        raise HTTPException(status_code=500, detail=error_msg)

@app.post("/api/excel/export")
async def export_workbook(
    request: dict,
    x_session_id: str = Header(None, alias="X-Session-Id")
):
    """Export workbook in specified format."""
    session_id = request.get("sessionId") or x_session_id
    session = get_session(session_id)
    workbook = session.get("workbook")

    if not workbook:
        raise HTTPException(status_code=400, detail="No workbook loaded")

    format_type = request.get("format", "xlsx")

    # Create temporary export file
    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{format_type}") as tmp:
        if format_type == "xlsx":
            workbook.save(tmp.name)
        else:
            # Add CSV export logic if needed
            workbook.save(tmp.name)

        # Return file download
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"exported_workbook.{format_type}"
        )

@app.get("/api/excel/stream/{session_id}")
async def stream_events(session_id: str):
    """Server-Sent Events endpoint for real-time updates."""

    async def event_generator():
        session = get_session(session_id)
        last_event_idx = 0

        yield f"data: {json.dumps({'type': 'connected', 'message': 'Stream connected'})}\n\n"

        while True:
            try:
                session = get_session(session_id)
                events = session.get("events", [])

                # Send new events
                for event in events[last_event_idx:]:
                    yield f"data: {event.json()}\n\n"

                last_event_idx = len(events)
                await asyncio.sleep(0.5)  # Poll every 500ms

            except HTTPException:
                # Session expired or not found
                yield f"data: {json.dumps({'type': 'error', 'message': 'Session expired'})}\n\n"
                break
            except Exception as e:
                logger.error(f"SSE error: {e}")
                break

    return StreamingResponse(
        event_generator(),
        media_type="text/plain",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Content-Type": "text/event-stream"
        }
    )

# === Background Tasks ===

@app.on_event("startup")
async def startup_event():
    """Clean up expired sessions on startup."""
    logger.info("Excel Agent API started")

@app.on_event("shutdown")
async def shutdown_event():
    """Clean up all sessions on shutdown."""
    for session_id in list(sessions.keys()):
        cleanup_session(session_id)
    logger.info("Excel Agent API stopped")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=3005)